from openpyxl import load_workbook
import gdown

def home(currentwork):
    f = open("test.html" , "r")
    lines = f.readlines()
    nlines = len(lines) #count the number of lines
    f = open("test.html" , "w")
    for i in range(0 , nlines-4):
        f.write(lines[i])              #cursor reaches to the line where we need to edit
    f.write("\n")
    a2=currentwork["A2"]
    f.write("<a href=\""+a2.value+".html\">\n")
    html = a2.value+".html"
    f2 = open(html , "w")
    f2.close()
    f.write("<div id=\"b3\">")
    f.write("<br class=\"hiden\">\n")
    f.write("<img src=\""+a2.value+".png\" width = \"100%\" >\n")
    e2 = currentwork["E2"]
    url = e2.value
    output = a2.value+".png"
    gdown.download(url, output)
    f.write("<h5>"+a2.value+"</h5>\n")
    b2 = currentwork["B2"]
    f.write("<span>"+b2.value+"</span>\n")
    c2 = currentwork["C2"]
    f.write("<div> &#x20b9; "+str(c2.value)+"</div><br class=\"hiden\"></div></a>\n")
    f.write(lines[-4])
    f.write(lines[-3])
    f.write(lines[-2])
    f.write(lines[-1])
    f.close()

def cat(currentwork):
    

def ind(currentwork):
    a2=currentwork["A2"]
    url = a2.value +".html"
    f = open(url , "w")
    f.write("<!doctype html>\n")
    f.write("<html>\n")
    f.write("<head>\n")
    f.write("<meta charset=\"utf-8\">\n")
    f.write("<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">\n")
    f.write("<title> Swarnakshara </title>\n")
    f.write("<link = rel =\"stylesheet\" href=\"style.css\">\n")
    f.write("<style>\n")
    f.write("*{ padding: 0px; margin : 0px; box-sizing: border-box; }\n")
    f.write("body{background: #FAFAFF; background-image: url(\"circle.svg\"); background-repeat: no-repeat; background-position: 340px 280px; background-attachment: fixed; padding: 5%; padding-bottom: 10%; font-size: larger;}\n")
    f.write("</style>\n")
    f.write("</head>\n")
    f.write("<body>\n")
    f.write("<a id =\"it\"  href = \"Home.html\"  id=\"\"><img src=\"backarrow.svg\"></a>\n")
    f.write("<div id=\"spaceforl\"><br><br><br></div>\n")
    f.write("<div id=\"out\"><img src=\""+a2.value+".png\" width=\"100%\"; ></div>\n")
    f.write("<div id =\"out1\">\n")
    f.write("<h1>" + a2.value+"</h1>\n")
    f.write("<br>\n")
    b2 = currentwork["B2"]
    f.write("<h5>"+b2.value+"</h5>\n")
    h2 = currentwork["H2"]
    f.write("<br>\n")
    f.write(h2.value)
    f.write("<br><br><br>\n")
    f.write("</div>\n")
    c2 = currentwork["C2"]
    f.write("<div id=\"cost\" style =\"font-size: 1.2em;\"> &#x20b9;" +str(c2.value) +"</div>\n")
    f.write("<a href = \"https://forms.gle/MANUz25Fq4J1yHhc9\" target=\"_blank\">\n")
    f.write("<div id=\"orderb\" >Book an order</div>\n")
    f.write("</a>\n")
    f.write("<div id = \"out2\">\n")
    f.write("<img src=\""+a1.value+"1.png\" width=\"230px\" >\n")
    f.write("<br>\n")
    f.write("<img src=\""+a1.value+"2.png\" width=\"230px\" >\n")
    f.write("</div>\n")
    f.write("<div id=\"blogo\" ><img src=\"logo.svg\" width=\"90%\"></div>\n")
    f.write("</body>\n")
    f.write("</html>\n")
    
workbook = load_workbook("wsheet.xlsx")
currentwork = workbook[workbook.sheetnames[0]]

while (currentwork["A2"].value != None ):
    home(currentwork)
    currentwork.delete_rows(2, 1)

workbook.save("wsheet.xlsx")


