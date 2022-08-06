from openpyxl import load_workbook

workbook = load_workbook("wsheet.xlsx")

currentwork = workbook[workbook.sheetnames[0]]

acell = currentwork["B2"]

print(acell.value)

currentwork.delete_rows(1, 1)

workbook.save("wsheet.xlsx")

